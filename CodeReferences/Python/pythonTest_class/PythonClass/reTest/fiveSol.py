import re
from statistics import mean, stdev
from collections import Counter
import openpyxl
from openpyxl.chart import BarChart,\
    LineChart,Reference

fp = open( 'reTest/localhost_access_log.txt')
data = []
def makeData():
    for line in fp:
        match= re.search('([\d.:]+) \S+ \S+ \[([\w/:]+) [\d+]+\] "[A-Z]+ (\S+) [\w./]+" \d+ ([\d-]+)' , line )
        ip = match.group(1)
        dt = match.group(2)
        view = match.group(3)
        bsize = match.group(4)
        bsize = 0 if bsize=='-' else int( bsize)
        data.append( (ip,dt,view,bsize) )

        """ Test for understanding """
        match_test = re.search('([\d.:]+) \S+ \S+ \[([\w/:]+)' , line )
        whole = match_test.group(0)
        ip = match_test.group(1)
        dt = match_test.group(2)

        match_test = re.search('([\d.:]+) \S+ \S+ \[([\w/:]+) [\d+]+\] "[A-Z]+' , line )
        whole = match_test.group(0)
        ip = match_test.group(1)
        dt = match_test.group(2)

        match_test = re.search('([\d.:]+) \S+ \S+ \[([\w/:]+) [\d+]+\] "[A-Z]+ (\S+)' , line )
        whole = match_test.group(0)
        ip = match_test.group(1)
        dt = match_test.group(2)
        view = match.group(3)

        match_test = re.search('([\d.:]+) \S+ \S+ \[([\w/:]+) [\d+]+\] "[A-Z]+ (\S+) [\w./]+"' , line )
        whole = match_test.group(0)
        ip = match_test.group(0)
        dt = match_test.group(1)
        view = match.group(3)

        match_test = re.search('([\w/:]+)' , line )
        whole = match_test.group(0)
        ip = match_test.group(1)

def showAll():
    for dt in data:
        print("ip:",dt[0] )
        print("dt:",dt[1] )
        print("view:",dt[2] )
        print("================")
def totBsize():
    print("전체사이즈:", sum( [n[3] for n in data] ) )
    print("전체평균:", mean( [n[3] for n in data] ) )
    print("전체편차:", stdev( [n[3] for n in data] ) )

def ipCntFn():
    ipList = [ n[0] for n in data]
    ipCount = Counter( ipList)
    for ip, cnt in ipCount.most_common():
        print("ip:",ip, '방문수:', cnt)
def viewCntFn():
    viewList = [ n[2] for n in data]
    viewCount = Counter( viewList)
    print( viewCount.most_common())
    for v, cnt in viewCount.most_common():
        print("뷰:", v, "방문수:", cnt)

def mostIPFn():
    ipList = [ n[0] for n in data]
    ipCount = Counter( ipList)
    ipCom = ipCount.most_common(1)[0]
    print( "ip:",ipCom[0], "방문수:", ipCom[1] )

def ipTop5Fn():
    ipList = [ n[0] for n in data]
    ipCount = Counter( ipList)
    ipCom = ipCount.most_common(5)
    for ip, cnt in ipCom:
        print("ip:",ip, "방문수:", cnt)

makeData()

wb = openpyxl.Workbook()
ws = wb.active
ws.append(['뷰','방문수'] )

viewList = [ n[2] for n in data]
viewCount = Counter( viewList)
for v, cnt in viewCount.most_common():
    ws.append( (v,cnt) )

chart = BarChart()
cdata = Reference( ws, min_col=2, min_row=1,max_row=ws.max_row)
cat = Reference( ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data( cdata, titles_from_data=True )
chart.set_categories( cat )
ws.add_chart( chart, 'F1' )
wb.save('reTest/view.xlsx')









