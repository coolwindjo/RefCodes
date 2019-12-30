#pip install openpyxl
import openpyxl
def writeXL():
    wb = openpyxl.Workbook()
    sh1 = wb.active #기본으로 생성되는 sheet 참조를 얻는다
    sh1.title = 'sheet1' # Sheet
    sh2 = wb.create_sheet('sheet2')
    sh1['A1'] = 10
    sh1['A2'] = 20
    sh1['A3'] ='=sum(A1:A2)'
    sh1['B1'] ='hello'
    sh2['A1'] ='python'
    wb.save('xlsTest/result/my.xlsx')
#read
wb = openpyxl.load_workbook('xlsTest/result/my.xlsx')
sh1 = wb['sheet1']
sh2 = wb['sheet2']
print(  sh1['A1'].value )
print(  sh2['A1'].value )
sh1['c1'] ='aaa'
wb.save('xlsTest/result/my.xlsx')










