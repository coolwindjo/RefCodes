import openpyxl
def xlWrite():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append( [10,20,30] )
    ws.append( [40,50,60] )
    ws.append( [70,80,90] )
    ws.append( [30,40,50] )
    wb.save('xlsTest/result/my1.xlsx')
wb = openpyxl.load_workbook('xlsTest/result/my1.xlsx') #,data_only=True)
ws = wb['Sheet']
print( ws.max_row )
print( ws.max_column )
for c1,c2,c3 in ws:
    print( c1.value, c2.value, c3.value )
print("=======")
for r in ws['A']:
    print( r.value )
print("=======")
for r in ws[2]:
    print( r.value )
print("=======")
for c1,c2 in ws['B2':'C4']:
    print(c1.value, c2.value)









