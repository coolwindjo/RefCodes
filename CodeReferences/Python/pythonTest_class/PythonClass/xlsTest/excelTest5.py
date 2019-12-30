import openpyxl
from openpyxl.chart import BarChart,LineChart,Reference
wb = openpyxl.Workbook()
ws = wb.active
ws.append( ('이름', '국어','영어') )
while True:
    name = input('이름:')
    kor = int( input('국어:') )
    eng = int( input('영어:') )
    ws.append( (name, kor,eng) )
    yn = input("계속입력(y/n):")
    if yn =='n':
        break
chart = BarChart()
chart.title= '성적데이터'
chart.x_axis.title = '이름'
chart.y_axis.title = '점수'
cdata = Reference( ws,min_col=2, max_col=3,
                   min_row=1, max_row=ws.max_row )
cat = Reference( ws,min_col=1,
                   min_row=2, max_row=ws.max_row )
chart.add_data(cdata,titles_from_data=True)
chart.set_categories( cat )
ws.add_chart( chart, 'F1')
wb.save('xlsTest/result/chart1.xlsx')




