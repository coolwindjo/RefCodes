import csv
import openpyxl
from openpyxl.chart import BarChart,Reference
data =[]
def makeData():
    fp = open('libTest/data/births.txt')
    rd = csv.reader(fp)
    for y,m,f in rd:
        data.append(( int(y), int(m), int(f) ) )
    fp.close()
makeData()
#1.남아수총합을 구하시요
print( sum( [n[1] for n in data] ) )
#2.남아수가 가장많은 년도와 남아수를 구하시요
mx = max( data, key=lambda v:v[1] )
print( mx ,mx[0], mx[1] )
#3. 2000년도 이후 남아수 여아수에 대해 바차트를
#엑셀에 그리시요
f2000 = filter( lambda v:v[0]>=2000, data)
wb = openpyxl.Workbook()
ws = wb.active
ws.append(('년도','남아수','여아수'))
for n in f2000:
    ws.append( n )
chart = BarChart()
cdata = Reference(ws,min_col=2, max_col=3,
                      min_row=1, max_row=ws.max_row )
cat = Reference(ws, min_col=1,
                    min_row=2,max_row=ws.max_row )
chart.add_data( cdata, titles_from_data=True )
chart.set_categories( cat )
ws.add_chart( chart, 'F1' )
wb.save('libTest/result/births.xlsx')




