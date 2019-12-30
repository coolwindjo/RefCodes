import re, openpyxl
from collections import Counter
from statistics import mean, stdev
from openpyxl.chart import BarChart,LineChart,Reference

def extractInfo(str):
    rst = None
    try:
        match = re.search('[0:]+1|([1-2]?[0-9]?[0-9].){3}[1-2]?[0-9]?[0-9]', str)
        ip = match.group()

        match = re.search('\[([\w/:]+)', str)
        date = match.group(1)

        match = re.search('GET (/\S*)', str)
        view_page = match.group(1)

        match = re.split(' ', str)
        byte_info = match[len(match) - 1].strip()
        byte_info = 0 if byte_info == '-' else int(byte_info)

        rst = (ip, date, view_page, byte_info)
    except Exception as err:
        print(err, str)

    return rst

def readFile():
    data = []
    with open('localhost_access_log.txt', 'r') as fp:
        rd = fp.readlines()
        for str in rd:
            info = extractInfo(str)
            if info != None:
                data.append(info)
    return data

def printAllInfo(data):
    print("1번")
    for tmp in data:
        print("ip: ", tmp[0])
        print("날짜: ", tmp[1])
        print("뷰페이지: ", tmp[2])
        print("==================================")

def printIPAccessNum(data):
    print("************************************************")
    print("2번")
    cnt = Counter(n[0] for n in data)
    for key in cnt.keys():
        print("{} : {}번".format(key, cnt[key]))

def printViewNum(data):
    print("************************************************")
    print("3번")
    views = Counter(n[2] for n in data)
    for key in views.keys():
        print("{} : {}번".format(key, views[key]))

def printByteInfo(data):
    print("************************************************")
    print("4번")
    bytes = [n[3] for n in data]
    print("전체전송사이즈: ", sum(bytes))
    print("평균: ", mean(bytes))
    print("표준편차: ", stdev(bytes))

def printMaxIPAccessInfo(data):
    print("************************************************")
    print("5번")
    cnt = Counter(n[0] for n in data)
    maxIpInfo = max(cnt, key=lambda key:cnt[key])
    print("{}: {}".format(maxIpInfo, cnt[maxIpInfo]))


def printTopIPAccessInfo(data):
    print("************************************************")
    print("6번")
    cnt = Counter(n[0] for n in data)
    sortedIpAry = sorted(cnt, key=lambda key:cnt[key], reverse=True)

    for ipInfo in sortedIpAry[0:5]:
        print("{}: {}".format(ipInfo, cnt[ipInfo]))


def outputViewsToExcel(ws, views):
    ws.append(("페이지", "뷰수"))
    for key in views:
        ws.append((key, views[key]))

    chart = LineChart()
    chart.width = 20  # 기본값 15
    chart.height = 15  # 기본값 7
    chart.title = '페이지별 뷰수 그래프'
    chart.x_axis.title = '페이지'
    chart.y_axis.title = '뷰수'
    chart.style = 11
    cdata = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cat = Reference(ws, min_col=1,
                  min_row=2, max_row=ws.max_row)
    chart.add_data(cdata, titles_from_data=True)
    chart.set_categories( cat )
    ws.add_chart(chart, 'F1')


def outputIPToExcel(ws, ipAry):
    ws.append(("ip", "방문수"))
    for key in ipAry:
        ws.append((key, ipAry[key]))

    chart = BarChart()
    chart.width = 20  # 기본값 15
    chart.height = 15  # 기본값 7
    chart.title = 'ip별 방문수 그래프'
    chart.x_axis.title = 'ip'
    chart.y_axis.title = '방문수'
    chart.style = 11
    cdata = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cat = Reference(ws, min_col=1,
                  min_row=2, max_row=ws.max_row)
    chart.add_data(cdata, titles_from_data=True)
    chart.set_categories( cat )
    ws.add_chart(chart, 'F1')


def outputToExcel(data):
    views = Counter(n[2] for n in data)
    ipAry = Counter(n[0] for n in data)

    wb = openpyxl.Workbook()
    viewSh = wb.active
    ipSh = wb.create_sheet('sheet2')

    outputViewsToExcel(viewSh, views)
    outputIPToExcel(ipSh, ipAry)

    wb.save('work.xlsx')

data = readFile()
printAllInfo(data)
printIPAccessNum(data)
printViewNum(data)
printByteInfo(data)
printMaxIPAccessInfo(data)
printTopIPAccessInfo(data)
outputToExcel(data)
