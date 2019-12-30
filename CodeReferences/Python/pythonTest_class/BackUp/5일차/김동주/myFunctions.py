import re
from myFunctions import *
from statistics import *
from collections import Counter
import openpyxl
from openpyxl.chart import BarChart, Reference


def readFile():
    try:
        with open('localhost_access_log.txt', 'r') as fp:
            return(fp.readlines())
    except Exception as error:
        print("file open error")


def printMenu():
    print("===================================================")
    print("1번. 전체 로그 보기 ")
    print("2번. IP별 방문 횟수")
    print("3번. 페이지 별 뷰수")
    print("4번. 전체 전송 바이트 사이즈와 평균, 표준편차")
    print("5번. 가장 많이 방문한 IP와 방문 수")
    print("6번. 가장 많이 방문한 IP와 방문 수 top 5")
    print("7번. 엑셀 그래프로 저장 (페이지 별 뷰수, IP 별 방문수)")
    print("0번. 종료")

    return int(input("메뉴를 선택하시오: "))


def entireLog():
    rd = readFile()

    for n in rd:
        ip = re.search('\d{1,3}[.]\d{1,3}[.]\d{1,3}[.]\d{1,3}|[\d:]+', n)
        date = re.search('\[([\w/]+)', n)
        vp = re.search('(GET|POST)\s([\S]+)', n)
        print("--------------------------------------------------")
        print("ip:", ip.group())
        print("날짜:", date.group(1))
        print("뷰페이지:", vp.group(2))


def tranSize():
    rd = readFile()
    tb = 0
    data = []  # 각 byte 정보
    for n in rd:
        b = re.search('(\d+$)|([-]$)', n)
        # print("byte:", b.group())
        ib = int(b.group()) if b.group() != '-' else 0
        data.append(ib)
        tb += ib

    print("\n===================================================")
    print("전체 전송 바이트 사이즈: ", tb)
    print("평균: %.2f" % mean(data))
    print("표준 편차: %.2f" % stdev(data))


def mostVisIp():
    rd = readFile()
    data = []

    for n in rd:
        ip = re.search('\d{1,3}[.]\d{1,3}[.]\d{1,3}[.]\d{1,3}|[\d:]+', n)
        data.append(ip.group())

    c = Counter(data)

    ip = c.most_common(1)[0][0]
    cnt = c.most_common(1)[0][1]

    print('가장 많이 방문한 IP: ', ip, ': ', cnt, '번')


def mostVisIpTop5():
    rd = readFile()
    data = []

    for n in rd:
        ip = re.search('\d{1,3}[.]\d{1,3}[.]\d{1,3}[.]\d{1,3}|[\d:]+', n)
        data.append(ip.group())

    c = Counter(data)
    for ip, cnt in c.most_common(5):
        print(ip, ': ', cnt, '번')


def saveExcel():
    print('7 called')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(('IP','방문수', '페이지', '방문수'))

    rd = readFile()
    dataI = []
    for n in rd:
        ip = re.search('\d{1,3}[.]\d{1,3}[.]\d{1,3}[.]\d{1,3}|[\d:]+', n)
        dataI.append(ip.group())

    data1 = []
    c = Counter(dataI)
    for ip, cntI in c.most_common():
        data1.append((ip, cntI))


    dataV = []
    for n in rd:
        vp = re.search('(GET|POST)\s([\S]+)', n)
        dataV.append(vp.group(2))
    data2 = []
    c = Counter(dataV)
    for vp, cntV in c.most_common():
        data2.append((vp, cntV))

    data = list(zip(data1, data2))


    for n, m in data:
        ip = n[0]
        cnti = n[1]
        vp = m[0]
        cntv = m[1]
        ws.append((ip, cnti, vp, cntv))


    chart = BarChart()
    chart.title = 'IP 별 방문수'
    chart.x_axis.title = 'IP'
    chart.y_axis.title = '방문수'
    chart.style = 12
    chart.width = 20  # 15 in default (cell의 개수)
    chart.height = 15  # 7 in default
    cdata = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cat = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # categories
    chart.add_data(cdata, titles_from_data=True)
    chart.set_categories(cat)
    ws.add_chart(chart, 'F1')


    chart = BarChart()
    chart.title = '페이지 별 방문수'
    chart.x_axis.title = '페이지'
    chart.y_axis.title = '방문수'
    chart.style = 12
    chart.width = 20
    chart.height = 15
    cdata = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    cat = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)  # categories
    chart.add_data(cdata, titles_from_data=True)
    chart.set_categories(cat)
    ws.add_chart(chart, 'F31')

    wb.save('finalTest.xlsx')



def visitNum():
    rd = readFile()
    data = []

    for n in rd:
        ip = re.search('\d{1,3}[.]\d{1,3}[.]\d{1,3}[.]\d{1,3}|[\d:]+', n)
        data.append(ip.group())

    c = Counter(data)
    for ip, cnt in c.most_common():
        print(ip, ': ', cnt, '번')


def viewNum():
    rd = readFile()
    data = []
    for n in rd:
        vp = re.search('(GET|POST)\s([\S]+)', n)
        data.append(vp.group(2))

    c = Counter(data)
    for vp, cnt in c.most_common():
        print('{:<45}: {}번'.format(vp, cnt))
