import re
import statistics
import openpyxl
from openpyxl.chart import LineChart,Reference

# 로그분석(정규식을 이용할것) 1번과 4번만 해도 기본점수이상
# 1. 전체 로그보기(ip: .. 날짜: .. 뷰페이지: )
# 2. ip별 방문 횟수( 000.000.000.000:x회)
# 3. 페이지별 뷰수(ex /a.htm:x회 )
# 4. 전체전송바이트 사이즈 와 평균, 표준편차
# 5. 가장많이 방문한 ip 와 방문수
# 6. 가장많이 방문한 ip 와 방문수 top 5
# 7. 엑셀 그래프로 저장(페이지별뷰수, ip별방문수 2개의 그래프)

class LogUnit:
    def __init__(self, ip, date, page, size):
        self.ip = ip
        self.date = date
        self.page = page
        self.size = int(size) if not size=='-' else 0

class LogAnalysis:
    def __init__(self):
        self.data = []
        self.visitCount = {}
        self.pageViewCount = {}
        self.loadLogFile('localhost_access_log.txt')

    def loadLogFile(self, path):
        fp = open(path,'r')
        line = 0
        for rd in fp:
            line += 1
            # localhost IPv6 IP = 0:0:0:0:0:0:0:1
            match = re.search('(^0:0:0:0:0:0:0:1)', rd ) or re.search('(^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', rd)
            ip = match.group(0)
            match = re.search('\[([\w/:\+0-9 ]+)\] \"([\w]+) (\S+) (\S+)\" (\d+) (\d+|-)', rd)
            unit = LogUnit(ip, match.group(1),  match.group(3), match.group(6))
            self.data.append(unit)
            self.visitCount[unit.ip] = self.visitCount.get(unit.ip, 0)+1
            self.pageViewCount[unit.page] = self.pageViewCount.get(unit.page, 0)+1
        fp.close()

    def getData(self):
        return self.data

    def getVisitCount(self):
        return self.visitCount

    def getPageViewCount(self):
        return self.pageViewCount

# 메뉴를 선택하시요:
# 1.번
# ip:
# 날짜:
# 뷰페이지: /aa.png
# ----------
# ..
def menu1(log):
    for d in log.getData():
        print("ip:", d.ip )
        print("날짜:", d.date )
        print("뷰페이지:", d.page)


# 2.번
# 0:0:0:0:0:0:0:1:n번
# 192.168.123.240:n번
# ..
def menu2(log):
    for k, v in log.getVisitCount().items():
        print(k, ":", v, "번")

# 3 번:
# aa.png: 번
# tomcat.png;X번
def menu3(log):
    for k, v in log.getPageViewCount().items():
        print(k, ":", v, "번")

# 4.번
# 전체전송사이즈:
def menu4(log):
    dataSize = []
    for d in log.getData():
        dataSize.append(d.size)

    print("사이즈:", sum(dataSize))
    print("평균:%.4f"%(statistics.mean(dataSize)))
    print("표준편차:%.4f"%(statistics.stdev(dataSize)))
    pass

# 5 번:
# 가장많이 방문ip: xxx.xxx.xxx.xxx
def menu5(log):
    item = max([n for n in log.getVisitCount().items()], key=lambda v:v[1])
    print(item[0], ":", item[1])

# 6 번:
# 가장많이 방문ip와 방문수 Top 5
def menu6(log):
    sorted_counts = sorted([n for n in log.getVisitCount().items()], key=lambda v:v[1], reverse=True)
    for d in sorted_counts[:5]:
        print(d[0], ":", d[1])

# 7번:
# 차트로 그리기(map plot lib)(페이지별뷰수, ip별방문수)
def menu7(log):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title= "IP별 방문수"
    ws2 = wb.create_sheet("페이지별 뷰수")
    ws1.append(['IP','방문수'] )
    for d in log.getVisitCount().items():
        ws1.append([d[0], d[1]])

    ws2.append(['페이지','방문수'] )
    for d in log.getPageViewCount().items():
        ws2.append([d[0], d[1]])

    chart = LineChart()
    chart.title= 'IP별 방문수'
    chart.x_axis.title = 'IP'
    chart.y_axis.title = '방문수'
    cdata = Reference(ws1,min_col=2,
                        min_row=1, max_row=ws1.max_row)
    chart.add_data(cdata,titles_from_data=True)
    ws1.add_chart( chart, 'F1')

    chart = LineChart()
    chart.title= '페이지별 방문수'
    chart.x_axis.title = '페이지'
    chart.y_axis.title = '방문수'
    cdata = Reference(ws2,min_col=2,
                        min_row=1, max_row=ws2.max_row)
    chart.add_data(cdata,titles_from_data=True)
    ws2.add_chart( chart, 'F1')
    wb.save("Quiz5-7_ViewCount.xlsx")


if __name__ == '__main__':
    menu = {1: menu1, 2: menu2, 3: menu3, 4: menu4, 5: menu5, 6: menu6, 7: menu7}
    log = LogAnalysis()
    while True:
        print("")
        print('1.전체로그 보기',
              '2.IP별 방문 횟수',
              '3.페이지별 뷰수',
              '4.전체전송바이트 사이즈와 평규, 표준편차',
              '5.가장 많이 방문한 IP와 방문수',
              '6.가장 많이 방문한 IP와 방문수 Top 5',
              '7.엑셀 그래프로 저장하기',
              '8.종료',
              sep='\n')
        nSel = int( input('번호를 입력하세요:') )
        if 1 <= nSel < len(menu):
            menu[nSel](log)
        elif nSel == 8:
            break;
        else:
            print("잘못된 메뉴입니다")
