import openpyxl
from openpyxl.chart import BarChart, Reference
from collections import Counter
import statistics
import re

ips = []
pages = []
bytes = []

wb = openpyxl.Workbook()
ws = wb.active

def displayMenu():
    print ('''
1. 전체 로그보기(ip: .. 날짜: .. 뷰페이지: )  
2. ip별 방문 횟수( 000.000.000.000:x회)
3. 페이지별 뷰수(ex a.htm:x회 )
4. 전체전송바이트 사이즈 와 평균, 표준편차
5. 가장많이 방문한 ip 와 방문수
6. 가장많이 방문한 ip 와 방문수 top 5
7. 엑셀 그래프로 저장(페이지별뷰수, ip별방문수 2개의 그래프)    
    ''')

def selectMenu():
    return int(input('메뉴를 선택하시요 : '))


def readIP():
    global ips
    global pages
    global bytes

    with open('localhost_access_log.txt', 'r') as fp:
        while True:
            rd = fp.readlines()
            if not rd:
                break
            for n in rd:
                try:
                    ip = re.search('(\d:\d:\d:\d:\d:\d:\d:\d|\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3})', n)
                    date = re.search('\[([\w/:]+)', n)
                    page = re.search('(GET |POST )(\S+)', n)
                    byte = re.search('(-$|\w+$)', n)


                    ips.append(ip.group(1))
                    pages.append(page.group(2))

                    # print (byte.group(1))

                    if byte.group(1)=='-':
                        bytes.append(0)
                    else:
                        bytes.append(int(byte.group(1)))

                except Exception as err:
                    print ('not found')
    # print (ips)
    # print (pages)
    # print (bytes)

# 1
def menu1():
    with open('localhost_access_log.txt', 'r') as fp:
        while True:
            rd = fp.readlines()
            if not rd:
                break
            for n in rd:
                try:
                    ip = re.search('(\d:\d:\d:\d:\d:\d:\d:\d|\d{1,3}.\d{1,3}.\d{1,3}.\d{1,3})', n)
                    date = re.search('\[([\w/:]+)', n)
                    page = re.search('(GET |POST )(\S+)', n)

                    print ('ip :', ip.group())
                    print ('date :', date.group(1))
                    print ('view page :', page.group(2))
                    print ('\n')

                except Exception as err:
                    print ('not found')

# 2
def menu2():
    global ips

    ips_ = list(set(ips))

    c = Counter(ips)

    for n in ips_:
        print (n, ' : ', c[n] , '번')
    # for ip, cnt in Counter(ips).most_common():
    #     print (ip, cnt)



# 3
def menu3():
    global pages

    pages_ = list(set(pages))

    c = Counter(pages)

    for n in pages_:
        print (n, ' : ', c[n] , '번')

# 4
def menu4():
    print ('전체전송사이즈 :', sum(int(n) for n in bytes))
    print ('평균 :', statistics.mean(bytes))
    print ('표준편차 :', statistics.stdev(bytes))
    pass

# 5
def menu5():
    global ips
    d = Counter(ips)
    for ip, cnt in d.most_common(1):
        print ('가장많이 방문ip: ', ip)

# 6
def menu6():
    global ips
    d = Counter(ips)
    i = 1
    for ip, cnt in d.most_common(5):
        print ('ip[top', i, ']: ', ip)
        i += 1

# 7
def menu7():
    chart = BarChart()
    chart.title = 'page score'
    chart.x_axis.title = 'Name'
    chart.y_axis.title = 'Score'

    #cdata = Reference(ws,range_string='Sheet!B1:C4')
    cdata = Reference(ws,min_col=2, max_col=3,
                            min_row=1, max_row=ws.max_row)
    #chart.add_data(cdata,titles_from_data=True)
    chart.add_data(cdata)

    cat = Reference(ws, min_col=1, max_col=1,
                            min_row=1, max_row=ws.max_row)
    chart.set_categories(cat)

    ws.add_chart( chart, 'F1')
    wb.save('student.xlsx')

def deffn():
    exit()

def main():

    readIP()
    listOfMenu = {1:menu1, 2:menu2, 3:menu3, 4:menu4, 5:menu5, 6:menu6, 7:menu7}

    while True:
        displayMenu()
        iMenu = selectMenu()

        listOfMenu.get(iMenu, deffn)()


if __name__ == '__main__':
    main()
