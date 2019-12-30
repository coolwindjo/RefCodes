import csv
import statistics
import openpyxl
from openpyxl.chart import LineChart,Reference
import sqlite3

fp = open('CCTV_in_Seoul.csv',encoding='utf-8')
rd = csv.reader(fp)

# 1. cctv 소계의 평균을 구하시요
def p1_sol():
    # print(rd)
    # for a,s,d,f,g,h in rd:
    #     print(a,s,d,f,g,h)

    f = filter(lambda v:v[1]!='소계', rd)
    print( statistics.mean( [ int(n[1]) for n in f] ) )
    fp.close()

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
def p2_sol():
    f = filter(lambda v:v[1]!='소계', rd)
    mx = max(f, key=lambda v:int(v[1]))
    print( 'max기관명:{}, max소계:{}'.format(mx[0],mx[1]) )
    fp.close()

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
def p3_sol():
    data = []
    f = filter(lambda v:v[1]!='소계', rd)
    for n in f:
        data.append(( n[0], int(n[1]) ))
    # print(list(data))
    data = sorted(data, key=lambda v:v[1], reverse=True)
    # print(list(data))
    count = 0
    for n in data:
        if count < 5:
            print( '{}순위 - 기관명:{}, 소계:{}'.format(count+1,n[0],n[1]) )
            count += 1
    fp.close()

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
def p4_sol():
    data = []
    f = filter(lambda v:v[1]!='소계', rd)
    for n in f:
        data.append(( n[0], int(n[1]) ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(('기관명','소계'))
    for n in data:
        ws.append( n )
    chart = LineChart()
    cdata = Reference(ws,min_col=1, max_col=2,
                          min_row=1, max_row=ws.max_row )
    cat = Reference(ws, min_col=1,
                        min_row=2,max_row=ws.max_row )
    chart.add_data( cdata, titles_from_data=True )
    chart.set_categories( cat )
    ws.add_chart( chart, 'F1' )
    wb.save('cctv.xlsx')

    fp.close()

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이
# 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
def p5_sol():
    data = []
    f = filter(lambda v:v[1]!='소계', rd)
    for n in f:
        data.append(( n[0], int(n[1]), float(n[2])/(float(n[3])+float(n[4])+float(n[5])) ))

    count = 0
    for n in data:
        print( 'cctv 증가율 - 기관명:{}, 증가율:{}'.format(n[0],n[2]) )
    fp.close()

# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
def p6_sol():
    data = []
    f = filter(lambda v:v[1]!='소계', rd)
    for n in f:
        data.append(( n[0], int(n[1]) ))

    try:
        sql = 'create table if not exists ' \
              'cctv(name varchar(20), count int)'
        db = sqlite3.connect('cctv.db')
        db.execute(sql)
        db.close()
        print('Create success')

    except Exception as err:
        print('에러',err)


    try:
        sql = "insert into cctv(name, count)" \
                  " values(?,?)"
        db = sqlite3.connect('cctv.db')

        for n in data:
            db.execute(sql, (n))
        db.commit()
        db.close()
        print('Insert success')

    except Exception as err:
        print('에러',err)

# p1_sol()
# p2_sol()
# p3_sol()
# p4_sol()
# p5_sol()
p6_sol()
