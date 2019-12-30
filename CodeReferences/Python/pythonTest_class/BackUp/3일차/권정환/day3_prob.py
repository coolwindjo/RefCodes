import csv
import sqlite3
import openpyxl
from openpyxl.chart import BarChart,LineChart,Reference

fp = open('CCTV_in_Seoul.csv', 'r', encoding='utf-8') #readmode의 경우 default이므로 ('my.csv') 로만 해도 됨.
rd = csv.reader(fp)
num_of_gu = 0
camera_data = []
camera_data_sort = []
for place, total, y3, y4, y5, y6 in rd:
    if num_of_gu > 0:
        camera_data.append((place, int(total), int(y3), int(y4) ,int(y5), int(y6)))
        camera_data_sort.append((place, int(total), int(y3), int(y4) ,int(y5), int(y6)))
    num_of_gu += 1
fp.close()

num_of_gu -= 1

# print(list(rd))

# prob 1
avg_camera = sum([n[1] for n in camera_data])/num_of_gu
print('1 : ', avg_camera)

# prob 2
max_camera = max(camera_data, key = lambda v:v[1])
print('2 : ', max_camera[0], max_camera[1])

# prob 3

sorted_data = sorted(camera_data_sort, key = lambda v:v[1], reverse=True)
print('3.')
print(sorted_data[0][0], sorted_data[0][1])
print(sorted_data[1][0], sorted_data[1][1])
print(sorted_data[2][0], sorted_data[2][1])
print(sorted_data[3][0], sorted_data[3][1])
print(sorted_data[4][0], sorted_data[4][1])

# prob 4
wb = openpyxl.Workbook()
ws = wb.active
ws.append(('기관명', '소계'))
for n in camera_data:
    ws.append((n[0], n[1]))

chart = LineChart()
chart.title= '카메라'
chart.x_axis.title = '구'
chart.y_axis.title = '대수'
cdata = Reference( ws,min_col=1, max_col=2,
                   min_row=1, max_row=ws.max_row )
cat = Reference( ws,min_col=1,
                   min_row=2, max_row=ws.max_row )
chart.add_data(cdata,titles_from_data=True)
chart.set_categories( cat )
ws.add_chart( chart, 'F1')
print('4 : excel make')
wb.save('day3_prob.xlsx')

# prob 5
print('5 : ')
print('기관명\t\tCCVT 증가율')
print('====================')
for n1, n2, n3, n4, n5, n6 in camera_data:
    print(n1, '\t\t', "%0.1f"%float(n3/(n4+n5+n6)))

# prob6
def createTable():
    try:
        sql = 'create table if not exists cctv(place varchar(20), ' \
        'total int, year_13 int, year_14 int, year_15 int, year_16 int)'
        db = sqlite3.connect('cctv.db')
        db.execute(sql)
        db.close()
        print("create succeed")
    except Exception as err:
        print("에러 :", err)

def insertTable(): # 여러 내용을 insert
    try:
        sql = "insert into cctv(place, total, year_13, year_14, year_15, year_16) values(?, ?, ?, ?, ?, ?)"
        db = sqlite3.connect('cctv.db')
        db.executemany(sql, camera_data)
        db.commit() # 명령 확정을 해야함!
        db.close()
        print("Insert succeed")
    except Exception as err:
        print("에러 :", err)

createTable()
insertTable()
