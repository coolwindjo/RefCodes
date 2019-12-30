#3일차 과제 정은혜

import sqlite3
import csv
from statistics import mean
import openpyxl
from openpyxl.chart import BarChart,LineChart,Reference

data =[]
def makeData():
    fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
    rd = csv.reader(fp)
    next(rd)
    for a,b,c,d,e,f in rd: #통계 분석을 위해 2차원 튜플로 변환
        data.append( ( a,int(b),int(c),int(d),int(e),int(f)  ) )
    fp.close()

# 1. cctv 소계의 평균을 구하시요
def MeanData():
    print('소계의 평균: ',mean([n[1] for n in data]))

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
def MaxData():
    mx = max(data, key=lambda v:v[1])
    print('소계가 가장많은 기관:', mx[0], ', 소계:',mx[1])

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
def Top5Data():
    Data = sorted(data, key=lambda v: v[1],reverse=True)[:5]
    for n in Data:
        print('기관명:', n[0],', 소계:', n[1])

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
def makeChart():
    wb= openpyxl.Workbook()
    ws =wb.active
    ws.append(('기관명', '소계'))
    for n in data:
        ws.append(n)
    chart = LineChart()
    chart.title= '기관별 cctv소계'
    cData = Reference(ws,min_col=1, max_col=2,min_row=1, max_row=ws.max_row )
    cat = Reference(ws,min_col=1, min_row=1, max_row=ws.max_row)
    chart.add_data(cData,titles_from_data=True)
    chart.set_categories( cat )
    ws.add_chart( chart, 'F1')
    wb.save('myCCtv.xlsx')

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이 출력하시요.
def Incrase():
    print('기관명 cctv증가율')
    print('======================')
    for n in data:
        print( n[0]  ,n[2]/(n[3]+n[4]+n[5]))


# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
def creatTable():
    try:
        sql='create table if not exists cctv(name varchar(20), count int)'
        db = sqlite3.connect('CCtv.db')
        db.execute(sql)
        db.close()
        print("create success")
    except Exception as err:
        print("에러",err)
def insertTable():
    try:
        sql="insert into cctv(name,count) values(?, ?)"
        db = sqlite3.connect('CCtv.db')
        dt =[]
        for n in data:
            dt.append( (n[0], n[1]))
        db.executemany(sql,dt)
        db.commit()
        db.close()
        print("insert success")
    except Exception as err:
        print("에러",err)



makeData()
print('1번문제')
MeanData()
print('2번문제')
MaxData()
print('3번문제')
Top5Data()
print('4번문제')
makeChart()
print('6번문제')
Incrase()
print('6번문제')
creatTable()
insertTable()
