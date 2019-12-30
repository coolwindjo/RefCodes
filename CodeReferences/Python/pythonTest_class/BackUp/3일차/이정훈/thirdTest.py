import csv
import statistics
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
import sqlite3


fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
rd = csv.reader(fp)

title = []
data = []

cnt = 0
for city, ea, y2013, y2014, y2015, y2016 in rd:
    if cnt == 0:
        cnt += 1
        title.append((city, ea, y2013, y2014, y2015, y2016))
    else:
        data.append((city, int(ea), int(y2013), int(y2014), int(y2015), int(y2016)))

fp.close()

# 1. cctv 소계의 평균을 구하시요
print ('1-------------------')
print (statistics.mean(n[1] for n in data))

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
print ('2-------------------')
temp = max(data, key=lambda v:v[1])
print (temp[0], temp[1])

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
print ('3-------------------')
data = sorted(data, key=lambda v:v[1], reverse=True)
i = 0
for n in data:
    i += 1
    if i > 5:
        break
    print (n[0], n[1])

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
print ('4-------------------')
wb = openpyxl.Workbook()
ws = wb.active

ws.append((title[0][0], title[0][1]))

for cty, ea, y2013, y2014, y2015, y2016 in data:
    ws.append((cty, ea))

chart = LineChart()

chart.x_axis.title = title[0][0]
chart.y_axis.title = title[0][1]

cdata = Reference(ws,min_col=2,
                        min_row=1, max_row=ws.max_row)
chart.add_data(cdata,titles_from_data=True)

cat = Reference(ws, min_col=1,
                        min_row=2, max_row=ws.max_row)
chart.set_categories(cat)

ws.add_chart( chart, 'F1')
wb.save('CCTV.xlsx')
print ('success to mask LineChart in CCTV.xlsx')

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이
print ('5-------------------')
# 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
#
# 기관명 cctv증가율
# ================
# 강남구  0.5
# 강동구  0.3
# ....

for cty, ea, y2013, y2014, y2015, y2016 in data:
    #print (cty, y2013, y2014, y2015, y2016 )
    print (cty, y2013/(y2014 + y2015 + y2016) )


# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
print ('6-------------------')
def createTable():
    try:
        sql = 'create table if not exists ' \
                'cctv(cty varchar(20), ea int)'

        # dbTest.db 접속, 없으면 생성
        db = sqlite3.connect('cctv.db')
        db.execute(sql)
        print ('success to create cctv.db')
        db.close()

    except Exception as err:
        print ('Error!!', err)

def insertTable(dt):
    try:
        sql = "insert into cctv(cty, ea) " \
              "values(?,?)"

        # dbTest.db 접속, 없으면 생성
        db = sqlite3.connect('cctv.db')
        db.executemany(sql, dt)
        db.commit()
        db.close()

        print ('success to insert data')
    except Exception as err:
        print ('Error!!', err)

createTable()

dt = []
for cty, ea, y2013, y2014, y2015, y2016 in data:
    dt.append((cty, ea))

insertTable(dt)

# m_num = 0
#
# for y, m, f in data:
#     m_num += int(m)
#
# print (m_num)
# print (sum( n[1] for n in data))
#
#
# year = 0
#
#
# mx = max(data, key=lambda v:v[1])
# print (mx, mx[0], mx[1], mx[2])
#
#
#
# wb = openpyxl.Workbook()
# ws = wb.active
#
# ws.append(('년도', '남자','여자'))
#
# # f2000 = filter(lambda v:v[0] >=2000, data)
# # for n in f2000
# #     ws.append(n)
#
# for y, m, f in data:
#     if 2000 <= int(y):
#         ws.append((y,m,f))
#
# chart = BarChart()
#
# chart.x_axis.title = 'sex'
# chart.y_axis.title = 'num'
#
# cdata = Reference(ws,min_col=2, max_col=3,
#                         min_row=1, max_row=ws.max_row)
# chart.add_data(cdata,titles_from_data=True)
#
# cat = Reference(ws, min_col=1,
#                         min_row=2, max_row=ws.max_row)
# chart.set_categories(cat)
#
# ws.add_chart( chart, 'F1')
# wb.save('birth.xlsx')
