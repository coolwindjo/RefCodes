import statistics
import csv
import openpyxl as ex
from openpyxl.chart import LineChart, Reference
import sqlite3

fp = open('CCTV_in_Seoul.csv',encoding='utf-8')
rd = csv.reader(fp)

data = []
header = ()
isFirst = True
for l in rd:
    if isFirst:
        isFirst = False
        header = (str(l[0]),l[1],l[2],l[3],l[4],l[5])
        continue
    data.append((l[0],int(l[1]),int(l[2]),int(l[3]),int(l[4]),int(l[5])))
    #data.append()
fp.close()

def mymax(dt, key):
    mx = None
    for n in dt:
        if mx==None or key(n)>key(mx):
            mx = n
    return mx
maxTuple = mymax(data, lambda v:v[1])
print('1. cctv 소계의 평균을 구하시오 : ',statistics.mean([n[1] for n in data]))
print('2. ccv 소계가 가장 많은 기관명, 소계를 구하시오 : ',maxTuple[0],maxTuple[1])
def top5(dt, key):
    tempDt = dt.copy()
    resultList = []
    i = 0
    while i<5:
        r = mymax(tempDt, key)
        tempDt.remove(r)
        resultList.append(r)
        i += 1
    return resultList

print('3. ccv 소계가 가장 많은 top5 기관명 소계를 구하시오 : ')
top5List = top5(data,lambda v:v[1])
for i in top5List:
    print(i[0],i[1])

wb = ex.Workbook()
ws = wb.active
ws.append(header)
for d in data:
    ws.append(d)

wb.save('cctv_chart.xlsx')

wb = ex.load_workbook('cctv_chart.xlsx')
ws = wb['Sheet']

chart = LineChart()
chart.title= 'cctv현황'
chart.x_axis.title = '기관명'
chart.y_axis.title = 'ccv소계'
chart.height = 15
chart.width = 20
cdata = Reference( ws,min_col=2, max_col=2,
                   min_row=1, max_row=ws.max_row )
cat = Reference( ws,min_col=1,
                    min_row=2, max_row=ws.max_row )
chart.add_data(cdata,titles_from_data=True)
chart.set_categories( cat )
ws.add_chart( chart, 'H1')
wb.save('cctv_chart.xlsx')

def per(item):
    return (item[0],item[2]/(item[3]+item[4]+item[5]))
print("기관명\t\tccv증가율")
print('-' * 30)
for n in data:
    기관명, 증가율 = per(n)
    print(기관명,증가율,sep='\t\t')

def create():
    try:
        sql = 'create table if not exists cctv(name varchar(20), sum int)'
        db= sqlite3.connect('cctv.db') #있으면 접속, 없으면 생성
        db.execute(sql)

        db.close()
        print('create success')
    except Exception as err:
        print("error",err)

def insertTable(list):
    try:
        sql = "insert into cctv(name, sum) values(?,?)"
        db = sqlite3.connect('cctv.db')  # 있으면 접속, 없으면 생성
        for i in list:
            db.execute(sql, (i[0], i[1]))
        db.commit()
        # db.execute(sql, (name, age, birth))
        db.close()
        print('insert success')
    except Exception as err:
        print("error", err)

create()
insertTable(data)