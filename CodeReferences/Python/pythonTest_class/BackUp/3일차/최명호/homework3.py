import csv
import openpyxl
import sqlite3
from statistics import mean
from openpyxl.chart import LineChart,Reference


fp = open('CCTV_in_Seoul.csv',encoding='utf-8') # txt 파일이지만 ,로 구분되기 때문에 csv 사용
rd = csv.reader(fp) # rd는 generator

data = []
next(rd)
for pos, total, before_2013, amount_2014, amount_2015, amount_2016 in rd:
    data.append((pos, int(total), int(before_2013), int(amount_2014), int(amount_2015), int(amount_2016))) # 통계 분석이 가능하도록 2차원 구조로...
# print(data)
fp.close()

# 1. cctv 소계의 평균을 구하시요
cctv_total_mean = mean( [ n[1] for n in data] )
print('================')
print('cctv 소계의 평균=',cctv_total_mean)
print('================')

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
cctv_total_max = max( data, key=lambda v:v[1])[1]
cctv_total_max_pos = max( data, key=lambda v:v[1])[0]
print('================')
print('cctv 소계가 가장많은 기관명=',cctv_total_max_pos,'으로 소계=',cctv_total_max)
print('================')

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
sorted_data = sorted(data, key=lambda v:v[1], reverse=True)[:5]
# print(sorted_data)
i = 1
print('================')
for n in sorted_data:
    print('ranking-',i,'로, 소계가 많은 기관명',n[0],'으로 소계',n[1])
    i += 1
print('================')

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
wb = openpyxl.Workbook()
ws = wb.active
# f = filter(lambda v:v[0]>=2000,data )
#"기관명","소계"
ws.append(['기관명','소계'])
for n in data:
     ws.append([n[0], n[1]])
# wb.save('cctv.xlsx')

wb = openpyxl.load_workbook('cctv.xlsx')
ws = wb['Sheet']
chart = LineChart()
chart.width = 20 #15
chart.height = 15 #10
chart.style = 10 #색상조합

chart.title ="CCTV 기관별 소계"

chart.x_axis.title = '기관명'
chart.y_axis.title ='소계'

# # ws.max_row, ws.max_column
chart_data = Reference(ws,min_col=2,
                        min_row=1, max_row=ws.max_row)

chart_cat = Reference(ws, min_col=1, max_col=1,
                      min_row=2, max_row=ws.max_row)

chart.add_data(chart_data, titles_from_data=True )
chart.set_categories( chart_cat )
ws.add_chart(chart, 'F1')

wb.save('cctv.xlsx')
wb.close()


# 5. 2013년도 대비 cctv증가율에 대해 다음과 같이
# 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
#
# 기관명 cctv증가율
# ================
# 강남구  0.5
# 강동구  0.3
# male_tot = sum( [ n[1] for n in data] )
print('기관명 cctv증가율')
print('================')
for n in data:
    print(n[0], n[2]/sum(n[2:]))
print('================')


# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
data_db =[]

def createTable():
    try:
        sql = 'create table if not exists cctv(name varchar(20), total_cctv_num int)'
        db = sqlite3.connect('cctv.db')
        db.execute(sql) # table 생성
        db.close()
        print('create table success')

    except Exception as err:    # error 발생 시 error 내용 확인을 위한 예외처리
        print('에러발생', err)

def insertTable2():
    try:
        sql = "insert into cctv(name, total_cctv_num) values(?, ?)"
        db = sqlite3.connect('cctv.db')
        # bulk로 data를 insert할수 있다.
        # data = [ ('임꺽정',40,'1989-04-12'), ('이이',50,'1989-05-12'), ('김철수',60,'1989-06-12'), ('박철남',70,'1989-07-12')]
        for n in data:
            data_db.append((n[0],n[1]))
        db.executemany(sql, data_db) # table 생성
        db.commit() # 명령 확정
        db.close()
        print('insert many table success')

    except Exception as err:    # error 발생 시 error 내용 확인을 위한 예외처리
        print('에러발생', err)

createTable()
insertTable2()
