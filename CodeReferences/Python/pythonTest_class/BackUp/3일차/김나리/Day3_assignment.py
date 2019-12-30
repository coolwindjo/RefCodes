import csv
from statistics import mean
import openpyxl
from openpyxl.chart import LineChart,Reference
import sqlite3

fp = open('CCTV_in_Seoul.csv',encoding='utf-8')
rd = csv.reader(fp)

item = []
data = []
cnt = 0
# for n in rd:
#     data.append(n)
for name,numNow,num2013,num2014,num2015,num2016 in rd:
    if cnt == 0:
        cnt += 1
        item.append((name,numNow,num2013,num2014,num2015,num2016))
    else:
        data.append((name,int(numNow),int(num2013),int(num2014),int(num2015),int(num2016)))
fp.close()

print(data)
cnt = 0
# 1. CCTV 소계 mean
CCTV_numNow_mean = mean( [n[1] for n in data if type(n[1])==int] )
print("CCTV 소계 mean: ",CCTV_numNow_mean)
# 2. CCTV 소계 max의 기관명 및 소계
CCTV_numNow_max = max(data, key=lambda v:v[1])
print("CCTV 소계 max의 기관명과 소계: %s, %d"%(CCTV_numNow_max[0], CCTV_numNow_max[1]))
# 3. CCTV 소계 가장 많은 top 5의 기관명 및 소계
data = sorted(data,key=lambda v:v[1],reverse=True)
print("CCTV 소계 Top5의 기관명과 소계: ")
for n in data:
    print("     %s   %d"%(n[0], n[1]))
    cnt += 1
    if cnt == 5:
        break
# 4. 기관별 CCTV 소계에 대한 라인차트를 엑셀에 그리시오
wb = openpyxl.Workbook()
ws = wb.active
ws.append(("기관명","소계","2013년도 이전","2014","2015","2016"))
for n in data:
    ws.append(n)

chart = LineChart()
chartData = Reference(ws, min_col=2,min_row=2,max_row=ws.max_row)
category = Reference(ws, min_col=1,min_row=1,max_row=ws.max_row)
chart.add_data(chartData,titles_from_data=True)
chart.set_categories(category)
chart.title = "기관 별 CCTV 소계"
chart.x_axis.title = "기관명"
chart.y_axis.title = "CCTV 소계"
ws.add_chart(chart, 'H1')
wb.save('CCTV_chart.xlsx')

#5. 2013년도 대비 CCTV 증가율을 기관 별로 출력하시오.
print("2013년도 대비 CCTV 증가율")
print("="*20)
for n in data:
    print("%s   %.1f"%(n[0], n[2]/(n[3]+n[4]+n[5])))

# 6. CCTV DB의 CCTV 테이블에 기관명 소계를 저장하시오.
# try:
#     sql = "create table if not exists" \
#           "CCTV(name varchar(20), numNow int)"
#     db = sqlite3.connect('CCTV.db')   # 데이터베이스 만들어주기
#     db.execute(sql)
#     db.close()
#     print("create success")
# except Exception as err:
#     print("에러",err)
#
# try:
#     sql = "insert into CCTV(name, num) " \
#           "values(?,?)"   # SQL에서는 "은 안씀 '만 쓰자!
#     db = sqlite3.connect('CCTV.db')
#     db.execute(sql, data)
#     db.commit()
#     db.close()
#     print("insert success")
# except Exception as err:
#     print("에러",err)
