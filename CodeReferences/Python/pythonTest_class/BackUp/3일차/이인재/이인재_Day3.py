import csv
from statistics import mean
import openpyxl
from openpyxl.chart import LineChart, Reference
import sqlite3

fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
rd = csv.reader(fp)
header = next(rd)
data =[]
for n in rd:
    data.append((n[0], int(n[1]),int(n[2]), int(n[3]), int(n[4]), int(n[5])))

# 1. cctv 소계의 평균을 구하시요
def CalculateTotalAvgCCTV():
    total_avg = mean([n[1] for n in data])
    print("소계평균=",total_avg)
    print("=============================================")

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
def FindMaxTotalCCTV():
    max_data = max(data, key=lambda v:v[1])
    print(max_data[0],max_data[1])
    print("=============================================")

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
def FindTop5TotalCCTV():
    sort_data_set = sorted(data, key=lambda v:v[1], reverse=True)
    sort_data_set = sort_data_set[:5]
    for n in sort_data_set:
        print(n[0], n[1])
    print("+++++++++++++++++++++++++++++++++++++++++++++")

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
def IncertLineChartToExcel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(('기관명', '소계'))
    for n in data:
        ws.append((n[0],n[1]))

    chart = LineChart()
    chart.title = '기관별 CCTV 소계'
    chart.x_axis.title = '기관명'
    chart.y_axis.title = '소계'

    cdata = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cat = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(cdata, titles_from_data=True)
    chart.set_categories(cat)
    ws.add_chart(chart, 'F1')
    wb.save('total_cctv_per_org.xlsx')

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이 출력하시요. (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
# 기관명 cctv증가율
# ================
# 강남구  0.5
# 강동구  0.3
# ....
def ShowIncrementRate():
    print('기관명\t\tCCTV증가율')
    print("=============================================")
    inc_rate=0.0
    for n in data:
        inc_rate = n[2]/(n[3]+n[4]+n[5])
        print(n[0],'\t',inc_rate)
print("+++++++++++++++++++++++++++++++++++++++++++++")

# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
def SaveCCTVDB():
    try:
        sql = 'create table if not exists cctv(org varchar(20), total_number int)'
        db = sqlite3.connect('cctv_org.db')
        db.execute(sql)
        print("db create success")
    except Exception as err:
        print("에러", err)

    try:
        sql="insert into cctv(org, total_number) values(?,?)"
        data_set=[(n[0],n[1]) for n in data]
        db.executemany(sql,data_set)
        db.commit()
        db.close()
        print("db insert success")
    except Exception as err:
        print("에러",err)

CalculateTotalAvgCCTV()
FindMaxTotalCCTV()
FindTop5TotalCCTV()
IncertLineChartToExcel()
ShowIncrementRate()
SaveCCTVDB()