import csv
import openpyxl
from openpyxl.chart import LineChart, Reference
import sqlite3
import statistics

# 1. cctv 소계의 평균을 구하시요
# "기관명","소계","2013년도 이전","2014년","2015년","2016년"
print("1)-----------------------------")
fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
rd = csv.reader(fp)
data = []
for area, total, cctv_b2014, cctv_2014, cctv_2015, cctv_2016 in rd:
    if total.isdigit():
        data.append((area, int(total), int(cctv_b2014), int(cctv_2014), int(cctv_2015), int(cctv_2016)))
fp.close()

print ("CCTV 소계의 평균:%.2f" % (statistics.mean(d[1] for d in data)))

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
print("2)-----------------------------")
max_cctv = max(data, key=lambda v:v[1])
print("CCTV 소계가 가장 많은 기관명, 소계:", max_cctv[0], max_cctv[1])

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
print("3)-----------------------------")
sorted_data = sorted(data, key=lambda v:v[1], reverse=True)
print("CCTV 소계가 가장많은 top5 기관")
for d in sorted_data[0:5]:
    print(d[0], d[1])

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
print("4)-----------------------------")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "cctv"
ws.append(["area", "total"])
for n in data:
    ws.append(n[0:2])

chart = LineChart()
chart.title= 'cctv chart'
chart.x_axis.title = '지역'
chart.y_axis.title = 'cctv'
cdata = Reference( ws, min_col=2, max_col=ws.max_column,
                   min_row=1, max_row=ws.max_row)
cat = Reference( ws, min_col=1,
                   min_row=2, max_row=ws.max_row)
chart.add_data(cdata, titles_from_data=True)
chart.set_categories( cat )
ws.add_chart( chart, 'F1')
wb.save("Quiz3_4.xlsx")
print("Quiz3_4.xlsx을 확인하세요")

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이
# 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
#
# 기관명 cctv증가율
# ================
# 강남구  0.5
# 강동구  0.3
# ....
print("5)-----------------------------")
print ("기관명 cctv증가율")
print ("================")
for n in data:
    print(n[0], "%.2f"%(n[2]/sum(n[3:6])))


# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
print("6)-----------------------------")
try:
    sql='create table if not exists ' \
        'areacctv(name varchar(20), total int)'
    db = sqlite3.connect('Quiz3_6.db')
    db.execute(sql)
    db.close()
except Exception as err:
    print("error:",err)

try:
    sql="insert into areacctv(name, total)" \
        " values(?,?)"
    db = sqlite3.connect('Quiz3_6.db')
    total_data =[n[0:2] for n in data]
    db.executemany(sql, total_data)
    db.commit()
    db.close()
except Exception as err:
    print("error",err)

print("Quiz3_6.db을 확인하세요")
