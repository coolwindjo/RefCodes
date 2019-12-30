import csv
import statistics
import openpyxl
from openpyxl.chart import BarChart, Reference, LineChart
import sqlite3

AllData = []
RegionData = []

fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
str = csv.reader(fp)
for line in str:
    AllData.append(line)

RegionData = AllData[1:]

print(AllData)
print(RegionData)

# 1. cctv 소계의 평균을 구하시요
temp_list = [int(n[1]) for n in RegionData]
# print(temp_list)
sol_avg = sum(temp_list) / len(temp_list)
print("sol1 = ", sol_avg)

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
mx = max(RegionData, key = lambda v:v[1])
print("sol2 = ","기관명:",mx[0], "소계:",mx[1])

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
RegionBackUpData = [(n[0], int(n[1])) for n in RegionData]
RegionBackUpData.sort(key=lambda v:v[1],reverse=True)

top5 = RegionBackUpData[0:5]
print(top5)

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요

RegionBackUpData = [(n[0], int(n[1])) for n in RegionData]

wb = openpyxl.Workbook()
ws = wb.active
ws.append(('기관명','소계'))
for n in RegionBackUpData:
    ws.append(n)
chart = LineChart()
cdata = Reference(ws,min_col=2, max_col=3,
                      min_row=1, max_row=ws.max_row )
cat = Reference(ws, min_col=1,
                    min_row=2,max_row=ws.max_row )
chart.add_data( cdata, titles_from_data=True )
chart.set_categories(cat)
ws.add_chart(chart, 'F1' )
wb.save('homework3.xlsx')

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이
# 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
RegionBackUpData = [(n[0], int(n[2]), int(n[3]),int(n[4]),int(n[5]) )for n in RegionData]
for rr in RegionBackUpData:
    temp = rr[1] / rr[2]+rr[3]+rr[4]
    print(rr[0], temp)

# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요

RegionBackUpData = [ (n[0], int(n[1])) for n in RegionData]

try:
    sql = 'create table if not exists student(name varchar(20), age int)'
    db = sqlite3.connect('homework.db')
    db.execute(sql)
    db.close()
    print("create success")
except Exception as err:
    print("error", err)

try:
    sql = "insert into student(name, age) values(?,?)"
    db = sqlite3.connect('homework.db')
    db.executemany(sql,RegionBackUpData)
    db.commit()
    db.close()
    print("insert success")
except Exception as err:
    print("error", err)
