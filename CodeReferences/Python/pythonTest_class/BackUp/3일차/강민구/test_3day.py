import csv
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from statistics import mean

data = []
fp = open('CCTV_in_Seoul.csv', encoding='utf-8') #미국에서 태어난 각 연도별 남아/여아 수
rd = csv.reader(fp)

for city, tot, a, b, c, d in rd:
    data.append((city, tot, a, b, c, d))
data.pop(0)
# print(data)
# 1. cctv 소계의 평균을 구하시요
meanA = mean([ int(n[1]) for n in data])
print(meanA)
# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
maxCity = max( data, key=lambda v:int(v[1]))
print(maxCity[0], maxCity[1])
# 3. cctv 소계가 가장많은 top5 기관명, 소계를 구하시요
sortData = sorted(data, key=lambda v:int(v[1]), reverse=True)
for dt in sortData:
    print(dt[0], dt[1])

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
wb = openpyxl.Workbook()
ws = wb.active
ws.append(("기관명","소계","2013년도 이전","2014년","2015년","2016년"))
for n in data:
    ws.append(n)

chart = LineChart()
chart.title = '기관별 cctv 소계'
chart.x_axis.title = 'cctv소계'
chart.y_axis.title = '기관명'
cdata = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=ws.max_row)
cat = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(cdata, titles_from_data=True)
chart.set_categories(cat)
ws.add_chart(chart, 'H1')   #F1셀에 차트를 추가
wb.save('cctv_tot.xlsx')

fp.close()
# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
