import csv, openpyxl, statistics
from openpyxl.chart import LineChart, Reference

def readData():
    data = []
    fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
    rd = csv.reader(fp)             # generator
    for d in rd:
        data.append(d)
    fp.close()
    data.pop(0)

    return data

def outputAvg(data):
    print("소계의 평균: ",statistics.mean(map(lambda v:int(v[1]), data)))

def outputMax(data):
    maxData = max(data, key=lambda v:int(v[1]))
    print("소계가 가장많은 기관명(소계): {}({})".format(maxData[0], maxData[1]))

def outputMaxFive(data):
    sortedData = sorted( data ,key=lambda v:int(v[1]), reverse=True)
    print("소계수가 top5인 기관명(소계)")
    for n in sortedData[:5]:
        print("{}({})".format(n[0], n[1]))

def outputToExcel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("기관", "소계"))
    for n in data:
        ws.append((n[0], int(n[1])))

    chart = LineChart()
    chart.width = 20  # 기본값 15
    chart.height = 15  # 기본값 7
    chart.title = 'CCTV'
    chart.x_axis.title = '기관명'
    chart.y_axis.title = '소계'
    chart.style = 11
    cdata = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cat = Reference(ws, min_col=1,
                  min_row=2, max_row=ws.max_row)
    chart.add_data(cdata, titles_from_data=True)
    chart.set_categories( cat )
    ws.add_chart(chart, 'F1')

    wb.save('work0724.xlsx')

def outputRate(data):
    print("기관명\tcctv증가율")
    print("===========================")
    for d in data:
        before = int(d[2])
        after = int(d[3]) + int(d[4]) + int(d[5])
        print(d[0], before / after)

data = readData()
outputAvg(data)
outputMax(data)
outputMaxFive(data)
outputToExcel(data)
outputRate(data)
