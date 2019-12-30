import csv
import sqlite3
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference

wb = openpyxl.Workbook()
ws = wb.active

fp = open('CCTV_in_Seoul.csv',encoding='utf-8', newline='')
rp = csv.reader(fp)

data =[]
data1 =[]
#csv.
#1. cctv 소계의 평균을 구하시요
#print(rp[1].max_row)
print("헤더:",  next(rp) )
for n in rp:
    ws.append((n[0],int(n[1]),int(n[2]),int(n[3]),int(n[4]),int(n[5])))
    data.append((n[0],int(n[1]),int(n[2]),int(n[3]),int(n[4]),int(n[5])))
    data1.append((n[0],int(n[2]),int(n[3]),int(n[4]),int(n[5])))
wb.save('tt.xlsx')

#1. cctv 소계의 평균을 구하시요
#print(sum([n[1].value for n in ws]) / ws.max_row)

#2. cctv 소계가 가장많은 기관명, 소계를 구하시요
#mx = max(data, key=lambda v:v[1])
#print(mx)

#3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
# cnt = 0
# for n in data:
#     mx = max(data, key=lambda v:v[1])
#     data.remove(mx)
#     print(mx)
#     cnt +=1
#     if cnt == 5:
#         break;

#4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
# chart = BarChart()
# chart.title= 'CCTV 소계'
# cat = Reference( ws,min_col=2,
#                    min_row=1, max_row=ws.max_row )
# chart.add_data(cat,titles_from_data=True)
# ws.add_chart( chart, 'K1')
# wb.save('tt.xlsx')

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이
# 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
# print('기관명 CCTV 증가율')
# for n in data:
#     res = n[2] / (n[3] + n[4] + n[5])
#     print(n[0] , round(res,1))

#6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
# try:
#     sql='create table if not exists cctv(name varchar(20), year1 int, year2 int, year3 int, year4 int)'
#     db = sqlite3.connect('test.db')
#     db.execute(sql)
#     db.close()
#     print("create success")
# except Exception as err:
#     print("에러",err)
#
# try:
#     sql = "insert into cctv(name, year1, year2, year3, year4) values(?,?,?,?,?)"
#     db = sqlite3.connect('test.db')
#     db.executemany(sql, data1)
#     db.commit()
#     db.close()
#     print('insert success')
# except Exception as err:
#     print("에러",err)






