import csv
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference

#1
fp = open('CCTV_in_Seoul.csv', 'r', newline='', encoding='utf-8')

data = []
rate = []
def makeData():
    rd = csv.reader(fp)
    next(rd)
    for a,b,c,d,e,f in rd :
        data.append( ( str(a), int(b), int(c), int(d), int(e), int(f) ) )


makeData()
print( sum( n[1] for n in data ) )

sr = sorted(data, key=lambda v:v[1], reverse=True)

count = 0
while count != 5:
    t1 = sr[count]
    count = count + 1
    print (t1)

mx = max(data, key=lambda v:v[1])
print (mx[0], mx[1])


wb = openpyxl.Workbook()
ws = wb.active

yn = ''

for a in data:
    ws.append( (a[0], a[1]) )

#wb.save('cctv.xlsx')

chart = LineChart()
# chart.title = '성적데이터'
# chart.width = 20 #default 15
# chart.height = 15 #defalt 7
chart.x_axis.title = '지역'
chart.y_axis.title = '소계'
# chart.style = 11

cdata = Reference(ws, range_string='Sheet!B1:B25')
#     cdata = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
cat = Reference(ws, range_string='Sheet!A1:A25')
#     cat = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(cdata, titles_from_data=False)
chart.set_categories(cat)
ws.add_chart( chart, 'F1')
wb.save('chart.xlsx')


print("기관명 CCTV 증가율")
print("=================")
for a in data:
    rate.append( (a[0], float(a[2]/(a[3] + a[4] + a[5])) ))
    for a in data:
        print ( a[0], float(a[2]/(a[3]+a[4]+a[5])) )







fp.close()


