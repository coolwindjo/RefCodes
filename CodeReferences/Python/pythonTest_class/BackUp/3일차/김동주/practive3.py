import csv
from statistics import mean
import openpyxl
from openpyxl.chart import LineChart, Reference
import sqlite3


def makeData(data):
    fp = open('CCTV_in_Seoul.csv', encoding='utf8')
    rd = csv.reader(fp)

    for org, subtotal, y13, y14, y15, y16 in rd:
        data.append((org, subtotal, y13, y14, y15, y16))
    fp.close()


data = []
makeData(data)

title = data[0]
data.remove(title)

content = []

for n in data:
    content.append((n[0], int(n[1]), int(n[2]), int(n[3]), int(n[4]), int(n[5])))



# 1. cctv 소계의 평균을 구하시요
print("=======================================")
print('1. cctv 소계의 평균을 구하시요')
print('cctv 소계의 평균: {}'.format(mean([n[1] for n in content])))

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
print("=======================================")
print('2. cctv 소계가 가장많은 기관명, 소계를 구하시요')
mSt = max(content, key=lambda v: v[1])
print('cctv 소계가 가장 많은 기관명: {}, 소계:{}'.format(mSt[0], mSt[1]))

# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
print("=======================================")
print("3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요")
temp = content

for n in range(5):
    m = max(temp, key=lambda v: v[1])
    print("{}위: {}".format(n+1, m[0]), end=' ')
    temp.remove(m)

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요

print("=======================================")
print("See practice3.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.append(title)
for n in content:
    ws.append(n)

chart = LineChart()
chart.title = '기관별 cctv 소계'
chart.x_axis.title = '지역'
chart.y_axis.title = '개수'
chart.style = 12
chart.width = 20
chart.height = 15
cdata = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
cat = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # categories
chart.add_data(cdata, titles_from_data=True)
chart.set_categories(cat)
ws.add_chart(chart, 'H1')

wb.save('practice3.xlsx')

# 5. 2013년도 대비 cctv증가율에 대해 다음 과 같이
# 출력하시요.
# (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
#
# 기관명 cctv증가율
# ================
# 강남구  0.5
# 강동구  0.3
# ....
print("=======================================")
print("5. 2013년도 대비 cctv증가율에 대해 다음 과 같이 출력하시요.")
print("{}\t\t{}".format('기관명', 'cctv 증가율'))
print("---------------------------------------")
for n in content:
    rst = n[2]/sum([n[3], n[4], n[5]])
    print(n[0], '\t\t%.2f'% float(rst))


# 6. cctv db 의 cctv 테이블에 기관명 소계를 저장하시요
def createTable():
    try:
        sql = 'create table if not exists cctv(org varchar(20), subtotal int)'
        db = sqlite3.connect('practice3.db')
        db.execute(sql)
        db.close()
        print("creation success")
    except Exception as err:
        print("에러", err)

def insertTable():
    try:
        sql = "insert into cctv(org, subtotal) values(?, ?)"
        db = sqlite3.connect('practice3.db')
        data = []
        for n in content:
            data.append((n[0], n[1]))
        db.executemany(sql, data)
        db.commit()
        db.close()
        print("insert success")
    except Exception as err:
        print("에러", err)

createTable()
insertTable()
