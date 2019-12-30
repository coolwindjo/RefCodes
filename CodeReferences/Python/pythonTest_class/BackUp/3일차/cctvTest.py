import csv
from statistics import mean
import openpyxl
from openpyxl.chart import LineChart, Reference
data =[]
def makeData():
    fp = open('CCTV_in_Seoul.csv',encoding='utf-8')
    rd = csv.reader(fp)
    title = next(rd)
    for n in rd:
        data.append( (n[0],int(n[1]),int(n[2]),
                      int(n[3]),int(n[4]),int( n[5]) ) )

# 1. cctv 소계의 평균을 구하시요
def totMean():
    print('소계평균',mean( [ n[1] for n in data] ) )

# 2. cctv 소계가 가장많은 기관명, 소계를 구하시요
def cctvMax():
    mx = max( data, key=lambda v:v[1])
    print(mx[0], mx[1])
# 3. cctv 소계가 가장많은 top5 기관명 소계를 구하시요
def top5Fn():
    top5 = sorted(data, key=lambda v:v[1], reverse=True)[:5]
    for n in top5:
        print(n[0], n[1])

# 4. 기관별 cctv소계에 대한 라인차트를 엑셀에 그리시요
def excelChart():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append( ['기관명','소계'] )
    for n in data:
        ws.append( [n[0],n[1]] )

    chart = LineChart()
    cdata = Reference( ws, min_col=2,
                       min_row=1, max_row=ws.max_row )
    cat = Reference( ws, min_col=1,
                     min_row=2,max_row=ws.max_row )
    chart.add_data(cdata, titles_from_data=True )
    chart.set_categories(cat)
    ws.add_chart(chart, 'F1')
    wb.save('cctv.xlsx')


makeData()




















