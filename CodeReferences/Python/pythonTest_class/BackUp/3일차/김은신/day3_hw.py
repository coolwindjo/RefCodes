import csv
import sqlite3
from statistics import mean
import openpyxl
from openpyxl.chart import LineChart,Reference

fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
rd = csv.reader(fp)
data = []
next(rd)
for n in rd:
    data.append((n[0], int(n[1]), int(n[2]), int(n[3]), int(n[4]), int(n[5])))
# fp.close()

#################################### 1
def printMean():
    print("CCTV 소계 평균: ", round(mean([n[1] for n in data]), 2))
    print("######################################################")
printMean()


#################################### 2
def printMax():
    mx = max(data, key=lambda v:v[1])
    print("CCTV 소계 많은 기관:", mx[0], ", 소계:", mx[1])
    print("######################################################")
printMax()


#################################### 3
def printTop5():
    sData = sorted(data, key=lambda v: v[1],reverse=True)[:5]
    cnt = 1
    for dt in sData:
        print(cnt, ". 기관명:", dt[0], ", 소계:", dt[1])
        cnt += 1
    print("######################################################")
printTop5()


#################################### 5
def printRate():
    print("\t기관명\t\tCCTV증가율")
    print("===================================")
    raeteData = []
    for n in data:
        raeteData.append((n[0], float(n[2]), float(n[3]), float(n[4]), float(n[5])))

    for dt in raeteData:
        print("\t", dt[0], "\t", float(dt[1] / (float(dt[2]) + float(dt[3]) + float(dt[4]))))
printRate()

#################################### 6
def createDB():
    try:
        sql = 'create table if not exists ' \
                  'cctv(name varchar(20), total int)'
        db = sqlite3.connect('cctv_total.db')
        db.execute(sql)
        db.close()
        print("create success")

    except Exception as err:
        print("Error", err)

def insertDB():
    try:
        sql = "insert into cctv(name, total) " \
              "values(?, ?)"
        db = sqlite3.connect('cctv_total.db')
        totalData = []
        for n in data:
            totalData.append(n[0], int(n[1]))
        db.executemany(sql, totalData)
        db.commit()
        db.close()
        print("insert success")

    except Exception as err:
        print("Error", err)
createDB()
insertDB()

#################################### 4
# def printLineChart():
wb = openpyxl.Workbook()
ws = wb.active
ws.append(['기관','소계'])
# global data
for n in data:
    ws.append([n[0], n[1]])

chart = LineChart()
chart.style = 10
chart.title ='CCTV소계'
chart.x_axis.title = "기관"
chart.y_axis.title = "소계"

data= Reference(ws, min_col=2,min_row=1,max_row=ws.max_row)
cat = Reference(ws, min_col=1,min_row=2,max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cat)
ws.add_chart(chart, 'F1')
wb.save('CCTV_total.xlsx')
# printLineChart()


fp.close()
